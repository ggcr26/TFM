import tkinter as tk
from tkinter import messagebox
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import traceback

#*******************IMPORTAR BASES DE DATOS***************
PAT = pd.read_excel('Datos de distribución.xlsx',sheet_name='PATOLOGOS')
TM = pd.read_excel('Datos de distribución.xlsx',sheet_name='MUESTRAS')
TM.fillna('--', inplace=True)

#Importa archivo externo con los datos de las variables de las checkbox y los slides. Separa cada linea y le elimnina el \n del final.
with open('Variables.txt') as f:
    Variables = f.readlines()
    for en,i in enumerate(Variables):
        Variables[en]=i.split(':')
        Variables[en][2]=float(Variables[en][2].replace('\n',''))
        

#********************FUNCIONES****************************
# Actualiza y guarda en un archivo externo llamado Variables, los valores de las checkbox y slides.
# Tambien actualiza los valores de las UCL corregidas 
def Actualiza_valor_wid(x=0):
    with open('Variables.txt', 'w') as f:
        for en,i in enumerate(dic_wid):
            f.write(i)
            f.write(':')
    #Guarda el valor de la variable donde se guarda el checkbox
            f.write(str(reg_wid[en].get()))     
            f.write(':')
    #Guarda el valor del slide
            f.write(str(dic_wid[i][1].get()))   
            f.write('\n')

    PAT['UCL CORREGIDAS']=PAT['UCL TOTALES']
    for en,i in enumerate(PAT['UCL CORREGIDAS']):
        PAT['UCL CORREGIDAS'][en]=PAT['UCL CORREGIDAS'][en]*(1/(Variables[en][2]/100))
    for en,i in enumerate(dic_wid):
        Variables[en][1]=reg_wid[en].get()
        Variables[en][2]=dic_wid[i][1].get()
    return


#Crea la funcion que genera el segundo desplegable
def tipo_desplegar(Especialidad_Var):
    #Definimos las opciones de tipo de muestra que concuerdan con la especialidad seleccionada
    Opciones2=TM['TIPO DE MUESTRA'][TM['ESPECIALIDAD']==Especialidad_Var]               
    #Creamos el desplegable
    Tipo_muestra_2_Menu=tk.OptionMenu(caja_textos, Tipo_muestra_elegida, *Opciones2)     
    Tipo_muestra_2_Menu.place(x=130, y=85)
    return 

#Selecciona al patologo. Primero actualiza los datos y crea una lista de patologos en formato panda con los nombres y estado de la checkbox.
#Despues crea una lista con lo que tienen la especialidad que coincide con la muestra seleccionada. y finalmente los va eliminando si tienen 0 en la checkbox
def Seleccion_patologo(x=0):
    Actualiza_valor_wid(x=0)
    chbx=[]
    nomb=[]
    #Creamos lista de patologos activos (en funcion del checkbox)
    for en,x in enumerate(Variables):                           
        nomb.append(x[0])
        chbx.append(int(x[1]))
    activos=pd.DataFrame({'NOMBRE':nomb,'activo':chbx})
    #Creamos variable mu con los datos de la muestra seleccionada
    mue=TM[TM['TIPO DE MUESTRA']==Tipo_muestra_elegida.get()] 
    #Le añade el ID de la muestra      
    mue['ID']=ID_Muestra.get()                          
    #Aisla la especialidad de la muestra      
    mue_espec=mue['ESPECIALIDAD'].iat[0]
    #Lista los patologos con esa especialidad                
    PAT_activo=PAT[PAT[mue_espec]=='x']
    #Elimina de la lista los que esten inactivos
    for en,i in enumerate(PAT_activo['NOMBRE']):        
        if int(activos['activo'][activos['NOMBRE']==i])==0:
            indice=PAT_activo.index[PAT_activo['NOMBRE']==i]
            PAT_activo = PAT_activo.drop(indice)
    #Elige al que menos UCL corregida tenga
    PAT_activo = PAT_activo.loc[PAT_activo['UCL CORREGIDAS'].idxmin()]    
    #Variable con el nombre del patologo
    Pat_selec=PAT_activo['NOMBRE']                                        
    Patologo_asignado.set(Pat_selec)
    #Exporta patologo seleccionado y muestra
    return (Pat_selec,mue)                                                

#Funcion que registra la muestra en un archivo externo, y las UCLS en el de la lista de patologos.
def Registra_UCLS_y_muestra():
    # Comprobamos que tengamos ID de la muestra. De lo contrario mostramos un mensaje de error:
    if Patologo_asignado.get()=='Seleccione o calcule un patólogo a quien asignar':
        tk.messagebox.showinfo(message='Por favor, seleccione un patólogo a quien asignar la muestra, o pulse "Calcular" para seleccionarlo un patólogo de forma automática.')
        return
    
    Reg_Muestra=TM[TM['TIPO DE MUESTRA']==Tipo_muestra_elegida.get()]                                               
    Reg_Muestra['ID']=ID_Muestra.get()                                                                           
    Reg_Muestra['PATOLOGO']=Patologo_asignado.get()                                                     
    Reg_mue_excel = openpyxl.load_workbook('Registro de muestras.xlsx')        
    hoja_Reg_excel = Reg_mue_excel.active                                                                   
    for r in dataframe_to_rows(Reg_Muestra, index=False, header=False):                                     
        hoja_Reg_excel.append(r)                                        
    Reg_mue_excel.save('Registro de muestras.xlsx')                            
    Reg_mue_excel.close()

    #Ahora registramos las UCLS en el patologo que corresponda
    Reg_UCL_PAT = openpyxl.load_workbook('Datos de distribución.xlsx')
    Reg_UCL_PAT.active
    hoja_activa = Reg_UCL_PAT['PATOLOGOS']
    UCL_original=PAT['UCL TOTALES'][PAT['NOMBRE']==Patologo_asignado.get()]
    UCL_nuevas=float(Reg_Muestra['UCL MICRO'])+float(UCL_original)
    #Identifica la posicion en la hoja excel de la casilla de las UCL del patologo indicado. +2 porque es la fila +2 del excel 
    posicion_patol_UCL=PAT.index[PAT['NOMBRE']==Patologo_asignado.get()][0]+2          
    posicion_columna=PAT.columns.get_loc('UCL TOTALES')+1

    hoja_activa.cell(row=posicion_patol_UCL, column=posicion_columna, value=float(UCL_nuevas))
    Reg_UCL_PAT.save('Datos de distribución.xlsx')
    Reg_UCL_PAT.close()
    #Mensaje de confirmación
    tk.messagebox.showinfo(message='Muestra:{id} asignada a {nombre}'.format(id=ID_Muestra.get(),nombre=Patologo_asignado.get()))
    return

# Funcion de manejo de errores
def reporte_error(*args):
    err = traceback.format_exception(*args)  
    for i in err:
        if "Permission denied: 'Datos de distribución.xlsx'" in i:
            messagebox.showerror(message='Por favor, cierre la base de datos de distribución')
            return
        if "Permission denied: 'Registro de muestras.xlsx'" in i:
            messagebox.showerror(message='Por favor, cierre la base de datos de muestras')
        if "ValueError: attempt to get argmin of an empty sequence" in i:
            messagebox.showerror(message='No hay patólogos disponibles para esta muestra.')
        if len(ID_Muestra.get())==0:
            tk.messagebox.showinfo(message='Por favor, introduzca identifiacación de la muestra.')
            return
        if Tipo_muestra_elegida.get()=='Selecciona muestra':
            tk.messagebox.showinfo(message='Por favor, introduzca un tipo de muestra.')
            return
        # else:
            # messagebox.showerror(message=err)

#Función del boton de resize de ventana
tam_vent=1
def resize():
    global tam_vent
    if tam_vent==1:
        ventana.geometry('520x{0}'.format(260+60+50*long+30))
        boton_resize.config(text="Contraer")
        tam_vent=0
    else:
        ventana.geometry('520x250')
        boton_resize.config(text="Opciones avanzadas")
        tam_vent=1
    return() 




#***********************GUI********************************************
#Ventana general
ventana = tk.Tk()      
long=len(PAT['NOMBRE'])
ventana.geometry('520x250')
ventana.title("Asignación de muestras")
ventana.config(bg='#2E4053')


    #CAJAS DE TEXTO:
caja_textos=tk.Frame(bg='#AEB6BF', height=125, width=500)

    #ID MUESTRA
tk.Label(caja_textos, text='ID de muestra:', font='Calibri 12', bg='#EAECEE').place(x=10, y=10)               
ID_Muestra=tk.Entry(caja_textos, bg='#EAECEE',font='Calibri 13')
ID_Muestra.place(x=130, y=10)                                                                               

    #TIPO MUESTRA
tk.Label(caja_textos, text='Tipo de muestra',font='Calibri 12', bg='#EAECEE').place(x=10, y=52)
#Variable en la que se guarda la especialidad seleccionada 
Especialidad_Var=tk.StringVar(value='Selecciona especialidad')      
#Opciones de especialidad para el desplegable
Opciones=TM['ESPECIALIDAD'].unique()                                
Especialidad_muestra_1=tk.OptionMenu(caja_textos,Especialidad_Var, *Opciones, command=lambda i:tipo_desplegar(Especialidad_Var.get()))   
#Localizacion desplegable
Especialidad_muestra_1.place(x=130, y=50)    
#Variable para seleccion de tipo de muestra resultado
Tipo_muestra_elegida=tk.StringVar(value='Selecciona muestra')                       

caja_textos.place(x=10, y=10)

    #BOTONES DE CALCULAR
botones_frame=tk.Frame(bg='#AEB6BF', height=50, width=500)

boton_calcular=tk.Button(botones_frame, text='Calcular', command=Seleccion_patologo)
boton_calcular.place(x=10, y=12)
boton_asignar=tk.Button(botones_frame, text='Asignar', command=Registra_UCLS_y_muestra)
boton_asignar.place(x=75, y=12)
Patologo_asignado=tk.StringVar(value='Seleccione o calcule un patólogo a quien asignar')
Desp_Patologo_asignado=tk.OptionMenu(botones_frame,Patologo_asignado, *PAT['NOMBRE'])   
Desp_Patologo_asignado.place(x=135, y=10)

botones_frame.place(x=10, y=150)

    #MOSTRAR OPCIONES AVANZADAS
boton_resize=tk.Button(ventana, text='Opciones avanzadas', command=resize)
boton_resize.place(x=10, y=210)     


    #FRAME QUE CONTENGA LOS WIDGETS DE LA CHECKBOX Y LOS SLIDES
caja_wid=tk.Frame(bg='#AEB6BF', height=60+50*long+10, width=500) 
caja_titulo_wid=tk.Frame(caja_wid, bg='#273746', height=35, width=480)
tk.Label(caja_titulo_wid, text='Nombre:', font='Calibri 12', bg='#EAECEE').place(x=5, y=5)  
tk.Label(caja_titulo_wid, text='Disponibilidad:', font='Calibri 12', bg='#EAECEE').place(x=500-210-100, y=5) 
tk.Label(caja_titulo_wid, text='% Carga laboral:', font='Calibri 12', bg='#EAECEE').place(x=500-140, y=5) 
#Creamos una lista con IntVar que serviran para guardar los valores de las checkbox. 
#El valor de IntVar es el que recuperamos del archivo Variables en elq ue se guardan los valores de cada variable
reg_wid=[]
for i in list(range(len(PAT['NOMBRE']))):
    reg_wid.append(tk.IntVar(value=int(Variables[i][1])))  

#Crea diccionario dic_wid con los widgets. [0]=Checkbox, [1]=Slide
dic_wid={}
for en,i in enumerate(PAT['NOMBRE']):
#Crea un Slide
    Slide=tk.Scale(caja_wid, orient='horizontal',from_=1, to=200,length=200, command=Actualiza_valor_wid, bg='#EAECEE')   
    Slide.set(int(Variables[en][2]))
#Crea boton
    Check_boton=tk.Checkbutton(caja_wid, variable=reg_wid[en], command=Actualiza_valor_wid, bg='#EAECEE')                
    dic_wid[i]=(Check_boton, Slide)
#Imprime el diccionario
for en,i in enumerate(dic_wid):
    #Nombres
    tk.Label(caja_wid, text=i, font='Calibri 12').place(x=10,y=50+20+en*50)        
    #Botones
    dic_wid[i][0].place(x=500-210-30,y=50+20+en*50)            
    #Slides       
    dic_wid[i][1].place(x=500-210,y=50+10+en*50)               

caja_wid.place(x=10, y=260)
caja_titulo_wid.place(x=10, y=10)

#mensajes de reporte de errores
ventana.report_callback_exception = reporte_error

ventana.mainloop()

